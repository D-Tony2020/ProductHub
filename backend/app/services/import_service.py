"""存量 Excel 导入：两阶段（dry-run → confirm）。

Excel 模板约定（表头第 1 行）：
- 固定列：root_type_code | price | currency | valid_from
- 属性列：``attr:<路径>``，路径 = 槽code.槽code.属性code（根属性直接写属性code），
  单元格填选项 code 或 label；
- 黑盒列：``part:<槽路径>``，单元格填 ``供应商名|件名``（自动建档/复用成品件）。

dry-run 把解析结果存入 import_batch.report_json；confirm 基于存档行在单事务内
全量重校验后入库（不信任 dry-run 时点，TOCTOU 安全）；file_hash 部分唯一索引
保证同一文件只能 committed 一次。
"""
import hashlib
import io
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models import (
    AttributeDef,
    AttributeOption,
    ComponentSlot,
    ImportBatch,
    NodeType,
    PurchasedPart,
    SkuPrice,
    Supplier,
)
from app.schemas.config import ConfigNodeIn, ConfigPayload, SlotSelection, AttributeSelection
from app.services.codes import next_code
from app.services.config_engine import create_sku, validate_config


class RowError(Exception):
    pass


def file_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _resolve_slot_path(db: Session, root_type: NodeType, path: list[str]) -> list[ComponentSlot]:
    """沿槽 code 路径解析出槽链。"""
    slots: list[ComponentSlot] = []
    current_type_id = root_type.id
    for code in path:
        slot = db.execute(
            select(ComponentSlot).where(
                ComponentSlot.parent_type_id == current_type_id,
                ComponentSlot.code == code,
            )
        ).scalar_one_or_none()
        if slot is None:
            raise RowError(f"槽路径不存在：{'.'.join(path)}（在 {code} 处中断）")
        slots.append(slot)
        current_type_id = slot.child_type_id
    return slots


def _resolve_option(db: Session, node_type_id: int, attr_code: str, raw: str) -> tuple[int, int]:
    attr = db.execute(
        select(AttributeDef).where(
            AttributeDef.node_type_id == node_type_id, AttributeDef.code == attr_code
        )
    ).scalar_one_or_none()
    if attr is None:
        raise RowError(f"属性不存在：{attr_code}")
    value = str(raw).strip()
    option = db.execute(
        select(AttributeOption).where(
            AttributeOption.attribute_id == attr.id,
            (AttributeOption.code == value.upper()) | (AttributeOption.label == value),
        )
    ).scalar_one_or_none()
    if option is None:
        raise RowError(f"属性「{attr.name}」没有匹配选项：{value}")
    return attr.id, option.id


def _resolve_part(
    db: Session, slot: ComponentSlot, raw: str, *, create_missing: bool, actor_id: int | None
) -> int:
    value = str(raw).strip()
    if "|" not in value:
        raise RowError(f"黑盒列格式应为「供应商名|件名」：{value}")
    supplier_name, part_name = (s.strip() for s in value.split("|", 1))
    part_name = " ".join(part_name.split())
    supplier = db.execute(
        select(Supplier).where(Supplier.name == supplier_name)
    ).scalar_one_or_none()
    part = None
    if supplier is not None:
        part = db.execute(
            select(PurchasedPart).where(
                PurchasedPart.supplier_id == supplier.id,
                PurchasedPart.node_type_id == slot.child_type_id,
                PurchasedPart.name == part_name,
            )
        ).scalar_one_or_none()
    if part is not None:
        if part.status not in ("draft", "active"):
            raise RowError(f"成品件「{part_name}」已合并或停用")
        return part.id
    if not create_missing:
        return -1  # dry-run 占位：确认阶段会创建
    if supplier is None:
        supplier = Supplier(code=next_code(db, "SUP"), name=supplier_name)
        db.add(supplier)
        db.flush()
    part = PurchasedPart(
        code=next_code(db, "PP"),
        node_type_id=slot.child_type_id,
        supplier_id=supplier.id,
        name=part_name,
        status="active",
        created_by=actor_id,
    )
    db.add(part)
    db.flush()
    return part.id


def _build_payload(
    db: Session, root_type: NodeType, attr_cells: dict[str, str], part_cells: dict[str, str],
    *, create_missing_parts: bool, actor_id: int | None,
) -> ConfigPayload:
    """由扁平列构建配置树。"""
    root = ConfigNodeIn()
    # path(tuple of slot codes) -> ConfigNodeIn
    nodes: dict[tuple[str, ...], ConfigNodeIn] = {(): root}
    slot_chains: dict[tuple[str, ...], list[ComponentSlot]] = {(): []}

    def ensure_node(path: tuple[str, ...]) -> ConfigNodeIn:
        if path in nodes:
            return nodes[path]
        parent = ensure_node(path[:-1])
        chain = _resolve_slot_path(db, root_type, list(path))
        slot_chains[path] = chain
        child = ConfigNodeIn()
        parent.slots.append(
            SlotSelection(slot_id=chain[-1].id, mode="configured", child=child)
        )
        nodes[path] = child
        return child

    for col_path, raw in attr_cells.items():
        parts = col_path.split(".")
        slot_path, attr_code = tuple(parts[:-1]), parts[-1]
        node = ensure_node(slot_path)
        if slot_path:
            node_type_id = slot_chains[slot_path][-1].child_type_id
        else:
            node_type_id = root_type.id
        attr_id, option_id = _resolve_option(db, node_type_id, attr_code, raw)
        node.attributes.append(AttributeSelection(attribute_id=attr_id, option_id=option_id))

    for col_path, raw in part_cells.items():
        slot_path = tuple(col_path.split("."))
        parent = ensure_node(slot_path[:-1])
        chain = _resolve_slot_path(db, root_type, list(slot_path))
        part_id = _resolve_part(
            db, chain[-1], raw, create_missing=create_missing_parts, actor_id=actor_id
        )
        parent.slots.append(
            SlotSelection(slot_id=chain[-1].id, mode="purchased", purchased_part_id=part_id)
        )

    return ConfigPayload(root_type_id=root_type.id, root=root)


def parse_workbook(content: bytes) -> tuple[list[str], list[dict]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RowError("Excel 为空")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for idx, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        out.append({"row_no": idx, "cells": cells})
    return headers, out


def process_row(
    db: Session, cells: dict, *, commit_mode: bool, actor_id: int | None,
    batch_id: int | None = None,
) -> dict:
    """处理一行：校验（与可选的落库）。返回报告项。commit_mode=False 时无副作用要求由
    外层 dry-run 事务整体回滚来保证（成品件占位不落库）。"""
    root_code = str(cells.get("root_type_code") or "").strip().upper()
    if not root_code:
        raise RowError("缺少 root_type_code")
    root_type = db.execute(
        select(NodeType).where(NodeType.code == root_code)
    ).scalar_one_or_none()
    if root_type is None:
        raise RowError(f"品类不存在：{root_code}")

    attr_cells, part_cells = {}, {}
    for header, value in cells.items():
        if value is None or str(value).strip() == "":
            continue
        if header.startswith("attr:"):
            attr_cells[header[5:].strip()] = value
        elif header.startswith("part:"):
            part_cells[header[5:].strip()] = value

    payload = _build_payload(
        db, root_type, attr_cells, part_cells,
        create_missing_parts=commit_mode, actor_id=actor_id,
    )
    if not commit_mode and any(
        s.purchased_part_id == -1
        for node in _iter_nodes(payload.root)
        for s in node.slots
        if s.mode == "purchased"
    ):
        # dry-run 中存在待新建成品件：无法算指纹，标记为"将新建件并入库"
        return {"status": "new", "note": "包含将自动建档的成品采购件", "payload": None}

    result, _ = validate_config(db, payload)
    if not result.complete:
        raise RowError("；".join(f"[{i.path}] {i.message}" for i in result.issues))

    price_raw = cells.get("price")
    price = None
    if price_raw is not None and str(price_raw).strip() != "":
        try:
            price = Decimal(str(price_raw))
        except InvalidOperation:
            raise RowError(f"价格非法：{price_raw}")
        if price < 0:
            raise RowError("价格不能为负")
    currency = str(cells.get("currency") or get_settings().default_currency).strip().upper()
    valid_from_raw = cells.get("valid_from")
    if valid_from_raw is None or str(valid_from_raw).strip() == "":
        valid_from = date.today()
    elif isinstance(valid_from_raw, date):
        valid_from = valid_from_raw
    else:
        valid_from = date.fromisoformat(str(valid_from_raw).strip()[:10])

    if not commit_mode:
        status = "exists" if result.matched_sku else "new"
        return {"status": status, "matched": result.matched_sku.sku_code if result.matched_sku else None}

    sku, created = create_sku(db, payload, created_by=actor_id, import_batch_id=batch_id)
    price_note = ""
    if price is not None:
        open_price = db.execute(
            select(SkuPrice).where(
                SkuPrice.sku_id == sku.id, SkuPrice.currency == currency,
                SkuPrice.valid_to.is_(None),
            )
        ).scalar_one_or_none()
        if open_price is None:
            db.add(SkuPrice(sku_id=sku.id, price=price, currency=currency,
                            valid_from=valid_from, created_by=actor_id,
                            note=f"导入批次 #{batch_id}"))
            price_note = "已录价"
        elif open_price.price == price:
            price_note = "价格一致跳过"
        else:
            price_note = f"已有现价 {open_price.price}，导入价 {price} 未覆盖（请人工处理）"
    return {"status": "created" if created else "exists",
            "sku_code": sku.sku_code, "price_note": price_note}


def _iter_nodes(node: ConfigNodeIn):
    yield node
    for s in node.slots:
        if s.child is not None:
            yield from _iter_nodes(s.child)
