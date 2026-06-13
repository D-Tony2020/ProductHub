"""报价单：多草稿并存、币种隔离、单价快照、导出前一致性校验、导出后冻结。"""
import io
from datetime import date, datetime
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.audit import write_audit
from app.core.config import get_settings
from app.core.db import get_db
from app.core.security import get_current_user
from app.models import AppUser, Quote, QuoteItem, Sku
from app.schemas.quote import (
    PriceCheckOut,
    QuoteIn,
    QuoteItemIn,
    QuoteItemOut,
    QuoteItemUpdate,
    QuoteOut,
    QuoteUpdate,
)
from app.services.codes import next_code
from app.services.config_engine import _current_prices
from app.services.health_engine import compute_health, load_sku_for_health
from app.services.template_service import get_or_404

router = APIRouter(prefix="/quotes", tags=["quotes"])


def _item_out(db: Session, item: QuoteItem, currency: str) -> QuoteItemOut:
    sku = db.get(Sku, item.sku_id)
    current = next((p.price for p in _current_prices(db, item.sku_id) if p.currency == currency), None)
    return QuoteItemOut(
        id=item.id, sku_id=item.sku_id,
        sku_code=sku.sku_code if sku else "", sku_name=sku.name if sku else "",
        qty=item.qty, snapshot_price=item.snapshot_price, unit_price=item.unit_price,
        line_total=item.unit_price * item.qty, line_note=item.line_note,
        price_changed=(current is not None and current != item.snapshot_price),
        current_price=current,
    )


def _quote_out(db: Session, quote: Quote) -> QuoteOut:
    items = [_item_out(db, i, quote.currency) for i in quote.items]
    return QuoteOut(
        id=quote.id, quote_no=quote.quote_no, customer_name=quote.customer_name,
        customer_contact=quote.customer_contact, currency=quote.currency,
        valid_until=quote.valid_until.isoformat() if quote.valid_until else None,
        notes=quote.notes, status=quote.status,
        created_at=quote.created_at.isoformat() if quote.created_at else None,
        exported_at=quote.exported_at.isoformat() if quote.exported_at else None,
        items=items, total=sum((i.line_total for i in items), Decimal("0")),
    )


def _editable(quote: Quote) -> None:
    if quote.status != "draft":
        raise HTTPException(409, "报价单已导出冻结，如需修改请复制为新草稿")


@router.get("", response_model=list[QuoteOut])
def list_quotes(
    mine: bool = True,
    db: Session = Depends(get_db),
    user: AppUser = Depends(get_current_user),
):
    stmt = select(Quote).order_by(Quote.id.desc()).limit(100)
    if mine:
        stmt = stmt.where(Quote.created_by == user.id)
    return [_quote_out(db, q) for q in db.execute(stmt).scalars().all()]


@router.post("", response_model=QuoteOut, status_code=201)
def create_quote(
    body: QuoteIn, db: Session = Depends(get_db), user: AppUser = Depends(get_current_user)
):
    quote = Quote(
        quote_no=next_code(db, "Q"),
        customer_name=body.customer_name,
        customer_contact=body.customer_contact,
        currency=body.currency or get_settings().default_currency,
        valid_until=date.fromisoformat(body.valid_until) if body.valid_until else None,
        notes=body.notes,
        created_by=user.id,
    )
    db.add(quote)
    db.flush()
    write_audit(db, actor_id=user.id, action="create", entity_type="quote",
                entity_id=quote.id, after={"quote_no": quote.quote_no})
    db.commit()
    return _quote_out(db, quote)


@router.get("/{quote_id}", response_model=QuoteOut)
def get_quote(
    quote_id: int, db: Session = Depends(get_db), _: AppUser = Depends(get_current_user)
):
    return _quote_out(db, get_or_404(db, Quote, quote_id))


@router.patch("/{quote_id}", response_model=QuoteOut)
def update_quote(
    quote_id: int, body: QuoteUpdate,
    db: Session = Depends(get_db), user: AppUser = Depends(get_current_user),
):
    quote = get_or_404(db, Quote, quote_id)
    _editable(quote)
    data = body.model_dump(exclude_unset=True)
    if "valid_until" in data and data["valid_until"]:
        data["valid_until"] = date.fromisoformat(data["valid_until"])
    for field, value in data.items():
        setattr(quote, field, value)
    db.commit()
    return _quote_out(db, quote)


@router.post("/{quote_id}/items", response_model=QuoteOut, status_code=201)
def add_item(
    quote_id: int, body: QuoteItemIn,
    db: Session = Depends(get_db), user: AppUser = Depends(get_current_user),
):
    quote = get_or_404(db, Quote, quote_id)
    _editable(quote)
    sku = get_or_404(db, Sku, body.sku_id)
    if sku.status != "active":
        raise HTTPException(409, f"SKU {sku.sku_code} 已作废，不可加入报价单")
    # 完整性硬闸（先于价格，是 SKU 本体缺陷）：缺必选/必配或违反互斥组的残货焊在货架内出不了门
    health = compute_health(db, load_sku_for_health(db, sku.id))
    if health.blocking:
        fams = health.families
        first = (fams.completeness or fams.structural)[0]
        raise HTTPException(409, detail={
            "code": "INCOMPLETE_SKU",
            "family": "completeness" if fams.completeness else "structural",
            "sku_code": sku.sku_code,
            "message": f"SKU {sku.sku_code} 配置不完整（{first.message}），不能加入报价单，请先治理",
            "issues": [{"path": i.path, "family": i.family, "message": i.message}
                       for i in fams.completeness + fams.structural],
        })
    current = next(
        (p.price for p in _current_prices(db, sku.id) if p.currency == quote.currency), None
    )
    if current is None:
        # 币种隔离 + 待录价拦截：没有该币种现价的 SKU 不能进单
        raise HTTPException(
            409,
            f"SKU {sku.sku_code} 没有 {quote.currency} 币种的现行价格"
            "（待录价或币种不符），不能加入本报价单",
        )
    existing = next((i for i in quote.items if i.sku_id == sku.id), None)
    if existing is not None:
        existing.qty += body.qty
    else:
        quote.items.append(
            QuoteItem(sku_id=sku.id, qty=body.qty, snapshot_price=current, unit_price=current)
        )
    db.commit()
    result = _quote_out(db, quote)
    # supply 软提醒（含停用/停产件）：仅当次加入/合并的那行回填，列表态不逐行重算
    if health.families.supply:
        warns = [i.message for i in health.families.supply]
        for it in result.items:
            if it.sku_id == sku.id:
                it.supply_warnings = warns
    return result


@router.patch("/{quote_id}/items/{item_id}", response_model=QuoteOut)
def update_item(
    quote_id: int, item_id: int, body: QuoteItemUpdate,
    db: Session = Depends(get_db), user: AppUser = Depends(get_current_user),
):
    quote = get_or_404(db, Quote, quote_id)
    _editable(quote)
    item = next((i for i in quote.items if i.id == item_id), None)
    if item is None:
        raise HTTPException(404, "明细行不存在")
    data = body.model_dump(exclude_unset=True)
    if "unit_price" in data and data["unit_price"] is not None:
        # 手动改价：保留快照价，差异留痕审计
        write_audit(db, actor_id=user.id, action="override_price", entity_type="quote_item",
                    entity_id=item.id,
                    before={"unit_price": str(item.unit_price)},
                    after={"unit_price": str(data["unit_price"]),
                           "snapshot_price": str(item.snapshot_price)})
    for field, value in data.items():
        if value is not None:
            setattr(item, field, value)
    db.commit()
    return _quote_out(db, quote)


@router.delete("/{quote_id}/items/{item_id}", response_model=QuoteOut)
def remove_item(
    quote_id: int, item_id: int,
    db: Session = Depends(get_db), _: AppUser = Depends(get_current_user),
):
    quote = get_or_404(db, Quote, quote_id)
    _editable(quote)
    item = next((i for i in quote.items if i.id == item_id), None)
    if item is None:
        raise HTTPException(404, "明细行不存在")
    quote.items.remove(item)
    db.commit()
    return _quote_out(db, quote)


@router.get("/{quote_id}/price-check", response_model=PriceCheckOut)
def price_check(
    quote_id: int, db: Session = Depends(get_db), _: AppUser = Depends(get_current_user)
):
    """导出前校验：快照价与现价是否一致。前端据此提示 [刷新为最新价]/[按原快照导出]。"""
    quote = get_or_404(db, Quote, quote_id)
    changed = [
        i for i in (_item_out(db, it, quote.currency) for it in quote.items) if i.price_changed
    ]
    return PriceCheckOut(consistent=not changed, changed_items=changed)


@router.post("/{quote_id}/refresh-prices", response_model=QuoteOut)
def refresh_prices(
    quote_id: int, db: Session = Depends(get_db), user: AppUser = Depends(get_current_user)
):
    quote = get_or_404(db, Quote, quote_id)
    _editable(quote)
    for item in quote.items:
        current = next(
            (p.price for p in _current_prices(db, item.sku_id) if p.currency == quote.currency),
            None,
        )
        if current is not None and current != item.snapshot_price:
            item.snapshot_price = current
            item.unit_price = current
    db.commit()
    return _quote_out(db, quote)


@router.post("/{quote_id}/duplicate", response_model=QuoteOut, status_code=201)
def duplicate(
    quote_id: int, db: Session = Depends(get_db), user: AppUser = Depends(get_current_user)
):
    src = get_or_404(db, Quote, quote_id)
    quote = Quote(
        quote_no=next_code(db, "Q"),
        customer_name=src.customer_name,
        customer_contact=src.customer_contact,
        currency=src.currency,
        valid_until=src.valid_until,
        notes=src.notes,
        created_by=user.id,
    )
    for item in src.items:
        quote.items.append(
            QuoteItem(sku_id=item.sku_id, qty=item.qty,
                      snapshot_price=item.snapshot_price, unit_price=item.unit_price,
                      line_note=item.line_note)
        )
    db.add(quote)
    db.flush()
    write_audit(db, actor_id=user.id, action="duplicate", entity_type="quote",
                entity_id=quote.id, note=f"复制自 {src.quote_no}")
    db.commit()
    return _quote_out(db, quote)


@router.post("/{quote_id}/export")
def export_excel(
    quote_id: int, force_snapshot: bool = False,
    db: Session = Depends(get_db), user: AppUser = Depends(get_current_user),
):
    """导出 Excel；导出后冻结（status=exported）。

    现价与快照不一致时默认拒绝，force_snapshot=true 表示业务员确认按原快照导出。
    """
    quote = get_or_404(db, Quote, quote_id)
    if not quote.items:
        raise HTTPException(409, "报价单没有明细行")
    if quote.status == "draft" and not force_snapshot:
        changed = [
            i for i in (_item_out(db, it, quote.currency) for it in quote.items)
            if i.price_changed
        ]
        if changed:
            raise HTTPException(
                409,
                f"{len(changed)} 个 SKU 价格已更新，请先刷新为最新价或确认按原快照导出"
                "（force_snapshot=true）",
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    bold = Font(bold=True, size=14)
    ws["A1"] = "报 价 单 / QUOTATION"
    ws["A1"].font = bold
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(["报价单号", quote.quote_no, "", "日期", datetime.now().strftime("%Y-%m-%d")])
    ws.append(["客户", quote.customer_name, "", "币种", quote.currency])
    if quote.valid_until:
        ws.append(["有效期至", quote.valid_until.isoformat()])
    ws.append([])
    header_row = ["#", "SKU 编码", "品名/规格", "数量", f"单价({quote.currency})", "小计"]
    ws.append(header_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    total = Decimal("0")
    for idx, item in enumerate(quote.items, start=1):
        sku = db.get(Sku, item.sku_id)
        line_total = item.unit_price * item.qty
        total += line_total
        ws.append([idx, sku.sku_code if sku else "", sku.name if sku else "",
                   item.qty, float(item.unit_price), float(line_total)])
    ws.append(["", "", "", "", "合计", float(total)])
    ws[ws.max_row][4].font = Font(bold=True)
    ws[ws.max_row][5].font = Font(bold=True)
    if quote.notes:
        ws.append([])
        ws.append(["备注", quote.notes])
    widths = [6, 18, 50, 8, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    if quote.status == "draft":
        quote.status = "exported"
        quote.exported_at = datetime.now()
        write_audit(db, actor_id=user.id, action="export", entity_type="quote",
                    entity_id=quote.id, after={"quote_no": quote.quote_no})
        db.commit()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{quote.quote_no}-{quote.customer_name}.xlsx"
    from urllib.parse import quote as urlquote

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{urlquote(filename)}"
        },
    )
